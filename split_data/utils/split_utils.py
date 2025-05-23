"""拆分工具类处理文件拆分核心逻辑"""

import os
import time
from math import ceil
from datetime import datetime
import pandas as pd
from multiprocessing import Pool
from openpyxl import load_workbook
import xlsxwriter
from tqdm import tqdm

from split_data.utils.file_utils import get_file_name
from split_data.utils.log_utils import log_info, log_error


class SplitUtils:
    """拆分工具类，提供静态方法进行文件拆分"""
    
    @staticmethod
    def write_csv_chunk(chunk, idx, input_file, output_folder):
        """写 CSV 每块：不再直接打印，而是返回信息字符串"""
        try:
            file_name = get_file_name(input_file, idx, "csv")
            path = os.path.join(output_folder, file_name)
            chunk.to_csv(path, index=False, encoding='utf-8-sig')
            msg = f"[CSV] 完成第 {idx} 块，共 {len(chunk)} 行 → {file_name}"
            log_info(f"[CSV] 已生成: {path} ({len(chunk)}行)")
            return msg
        except Exception as e:
            err_msg = f"[CSV] 第 {idx} 块写入失败: {e}"
            log_error(f"[CSV] 写入失败: {e}")
            return err_msg

    @staticmethod
    def write_excel_chunk(header, rows, idx, input_file, output_folder):
        """写 Excel 每块，同样返回信息字符串"""
        try:
            file_name = get_file_name(input_file, idx, "xlsx")
            path = os.path.join(output_folder, file_name)
            with xlsxwriter.Workbook(path, {'constant_memory': True}) as wb:
                ws = wb.add_worksheet()
                ws.write_row(0, 0, header)
                for r, row in enumerate(rows, start=1):
                    ws.write_row(r, 0, row)
            msg = f"[Excel] 完成第 {idx} 块，共 {len(rows)} 行 → {file_name}"
            log_info(f"[Excel] 已生成: {path} ({len(rows)}行)")
            return msg
        except Exception as e:
            err_msg = f"[Excel] 第 {idx} 块写入失败: {e}"
            log_error(f"[Excel] 写入失败: {e}")
            return err_msg


# 保留原函数，调用类方法
def write_csv_chunk(chunk, idx, input_file, output_folder):
    return SplitUtils.write_csv_chunk(chunk, idx, input_file, output_folder)


def write_excel_chunk(header, rows, idx, input_file, output_folder):
    return SplitUtils.write_excel_chunk(header, rows, idx, input_file, output_folder)


def split_csv_file(input_file, batch_size, output_folder, max_workers):
    """拆分 CSV 主流程"""
    log_info(f"开始拆分 CSV：batch_size={batch_size}, 并行度={max_workers}")

    # 统计总行数（减去表头）
    with open(input_file, 'r', encoding='utf-8-sig', errors='ignore') as f:
        total_rows = sum(1 for _ in f) - 1
    if total_rows <= 0:
        raise ValueError("CSV 文件为空或仅有表头")

    total_chunks = ceil(total_rows / batch_size)
    tqdm.write(f"CSV 总行数: {total_rows}, 预计拆分成 {total_chunks} 块，每块最多 {batch_size} 行")

    reader = pd.read_csv(input_file, chunksize=batch_size, low_memory=False)

    # 创建进度条, 将更新与信息打印都集中由主进程处理
    progress_bar = tqdm(total=total_chunks, desc="CSV拆分进度", unit="块", ncols=100)

    def update_progress(result):
        # 使用 tqdm.write 确保消息不干扰进度条
        tqdm.write(result)
        progress_bar.update(1)

    with Pool(processes=max_workers) as pool:
        for i, chunk in enumerate(reader, start=1):
            pool.apply_async(write_csv_chunk, args=(chunk, i, input_file, output_folder),
                            callback=update_progress)
        pool.close()
        pool.join()
    progress_bar.close()
    tqdm.write("[CSV] 所有块已处理完成。")
    return total_chunks


def split_excel_file(input_file, batch_size, output_folder, max_workers):
    """拆分 Excel 主流程"""
    log_info(f"开始拆分 Excel：batch_size={batch_size}, 并行度={max_workers}")

    wb = load_workbook(input_file, read_only=True)
    ws = wb.active

    total_rows = ws.max_row - 1
    if total_rows <= 0:
        raise ValueError("Excel 文件无数据内容")

    total_chunks = ceil(total_rows / batch_size)
    tqdm.write(f"Excel 总行数: {total_rows}, 预计拆分成 {total_chunks} 块，每块最多 {batch_size} 行")

    rows_gen = ws.iter_rows(values_only=True)
    try:
        header = next(rows_gen)
        if not any(header):
            raise ValueError("表头为空或格式异常")
    except Exception as e:
        raise ValueError(f"读取表头失败：{e}")

    batch = []
    idx = 1

    progress_bar = tqdm(total=total_chunks, desc="Excel拆分进度", unit="块", ncols=100)

    def update_progress(result):
        tqdm.write(result)
        progress_bar.update(1)

    with Pool(processes=max_workers) as pool:
        for row in rows_gen:
            batch.append(row)
            if len(batch) >= batch_size:
                pool.apply_async(write_excel_chunk, args=(header, batch.copy(), idx, input_file, output_folder),
                                callback=update_progress)
                idx += 1
                batch.clear()
        if batch:
            pool.apply_async(write_excel_chunk, args=(header, batch, idx, input_file, output_folder),
                            callback=update_progress)
        pool.close()
        pool.join()
    progress_bar.close()
    tqdm.write("[Excel] 所有块已处理完成。")
    return total_chunks
