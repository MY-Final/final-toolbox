"""
文件处理工具类
处理文件的读取和写入
"""
import os
import chardet
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import math


class FileUtils:
    """文件处理工具类，提供静态方法处理文件操作"""
    
    @staticmethod
    def get_file_extension(file_path):
        """获取文件扩展名"""
        return os.path.splitext(file_path)[1].lower()

    @staticmethod
    def create_output_folder(folder_name="拆分结果", clear_old=False):
        """创建输出文件夹，支持清空旧文件"""
        if clear_old and os.path.exists(folder_name):
            for f in os.listdir(folder_name):
                os.remove(os.path.join(folder_name, f))
        os.makedirs(folder_name, exist_ok=True)
        return folder_name

    @staticmethod
    def get_file_name(input_file, idx, ext):
        """获取拆分后文件命名（含日期和编号）"""
        base_name = os.path.splitext(os.path.basename(input_file))[0]
        date_prefix = datetime.now().strftime("%Y%m%d")
        return f"{base_name}_{date_prefix}_{idx:04d}.{ext}"

    @staticmethod
    def detect_encoding(file_path):
        """检测文件编码，更可靠的方法"""
        # 读取更多内容来提高检测准确性
        with open(file_path, 'rb') as f:
            raw_data = f.read(100000)  # 读取更多字节以提高准确性
            result = chardet.detect(raw_data)
        
        # 检查置信度和编码类型
        if result['confidence'] > 0.7:
            encoding = result['encoding']
        else:
            # 尝试常用编码列表
            encoding = 'utf-8'  # 默认尝试UTF-8
        
        # 对于中文环境，特别处理一些常见编码
        if encoding and encoding.lower() in ('gb2312', 'gbk', 'gb18030'):
            # 使用更兼容的GB18030替代GB2312
            return 'gb18030'
        elif encoding and encoding.lower() == 'ascii':
            # ASCII通常可以用UTF-8替代
            return 'utf-8'
        
        return encoding or 'utf-8'  # 如果检测失败则默认UTF-8

    @staticmethod
    def count_csv_rows(file_path):
        """计算CSV文件的总行数"""
        # 尝试多种编码
        encodings_to_try = ['utf-8', 'gb18030', 'utf-8-sig', 'latin1']
        
        # 先检测文件编码
        detected_encoding = FileUtils.detect_encoding(file_path)
        if detected_encoding and detected_encoding not in encodings_to_try:
            encodings_to_try.insert(0, detected_encoding)
        
        # 尝试所有编码
        for encoding in encodings_to_try:
            try:
                with open(file_path, 'r', encoding=encoding, errors='replace') as f:
                    total_rows = sum(1 for _ in f) - 1  # 减去表头
                return total_rows
            except Exception:
                continue
        
        # 如果所有编码都失败，使用二进制模式计算行数
        try:
            with open(file_path, 'rb') as f:
                total_rows = 0
                for _ in f:
                    total_rows += 1
            return total_rows - 1  # 减去表头
        except Exception as e:
            raise ValueError(f"无法读取CSV文件: {str(e)}")

    @staticmethod
    def read_csv_chunks(file_path, batch_size):
        """读取CSV文件，按块返回，增强编码处理"""
        # 尝试多种编码
        encodings_to_try = ['utf-8', 'gb18030', 'utf-8-sig', 'latin1']
        
        # 先检测文件编码
        detected_encoding = FileUtils.detect_encoding(file_path)
        if detected_encoding and detected_encoding not in encodings_to_try:
            encodings_to_try.insert(0, detected_encoding)
        
        # 尝试所有编码 - 使用C引擎
        for encoding in encodings_to_try:
            try:
                return pd.read_csv(
                    file_path, 
                    chunksize=batch_size, 
                    encoding=encoding, 
                    low_memory=False, 
                    on_bad_lines='skip'
                )
            except Exception:
                continue
        
        # 如果C引擎失败，尝试Python引擎 (不使用low_memory参数)
        for encoding in encodings_to_try:
            try:
                return pd.read_csv(
                    file_path, 
                    chunksize=batch_size, 
                    encoding=encoding, 
                    on_bad_lines='skip',
                    engine='python'  # 使用Python引擎增强兼容性
                )
            except Exception:
                continue
        
        # 最后尝试使用替换无法解析的字符
        try:
            return pd.read_csv(
                file_path, 
                chunksize=batch_size, 
                encoding='utf-8', 
                on_bad_lines='skip',
                engine='python',
                encoding_errors='replace'  # 替换无法解析的字符
            )
        except Exception as e:
            raise ValueError(f"无法读取CSV文件: {str(e)}")

    @staticmethod
    def get_excel_data(file_path):
        """获取Excel工作表和头部信息"""
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            
            # 检查是否有数据
            if ws.max_row <= 1:  # 只有表头或空表
                raise ValueError("Excel 文件无数据内容或只有表头")
                
            total_rows = ws.max_row - 1
            rows_gen = ws.iter_rows(values_only=True)
            header = next(rows_gen)
            
            # 验证表头是否有效
            if not any(header):
                raise ValueError("Excel 表头为空或无效")
                
            return ws, rows_gen, header, total_rows
        except Exception as e:
            if "Excel 文件无数据内容" not in str(e):
                raise ValueError(f"读取Excel文件失败: {str(e)}")
            raise

    @staticmethod
    def is_valid_file(file_path):
        """验证文件是否存在且格式受支持"""
        if not os.path.isfile(file_path):
            return False
            
        ext = FileUtils.get_file_extension(file_path)
        return ext in ['.csv', '.xlsx', '.xls']

    @staticmethod
    def get_file_info(file_path):
        """获取文件信息，包括类型、大小、行数等"""
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
            
        file_size = os.path.getsize(file_path)
        # 格式化文件大小
        if file_size < 1024:
            size_str = f"{file_size} 字节"
        elif file_size < 1024 * 1024:
            size_str = f"{file_size/1024:.2f} KB"
        else:
            size_str = f"{file_size/(1024*1024):.2f} MB"
        
        ext = FileUtils.get_file_extension(file_path)
        
        # 获取行数和类型信息
        if ext == '.csv':
            file_type = "CSV 文件"
            try:
                rows = FileUtils.count_csv_rows(file_path)
            except Exception as e:
                raise ValueError(f"读取CSV文件失败: {str(e)}")
        elif ext in ['.xlsx', '.xls']:
            file_type = "Excel 文件"
            try:
                wb = load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                if ws.max_row <= 1:
                    rows = 0
                else:
                    rows = ws.max_row - 1  # 减去表头
            except Exception as e:
                raise ValueError(f"读取Excel文件失败: {str(e)}")
        else:
            raise ValueError(f"不支持的文件格式: {ext}")
            
        # 默认每块最大行数
        batch_size = 49998
        chunks = math.ceil(rows / batch_size) if rows > 0 else 0
        
        return {
            "type": file_type,
            "size": size_str,
            "rows": rows,
            "chunks": chunks
        }


# 为了保持兼容性，提供直接的函数调用
def get_file_extension(file_path):
    return FileUtils.get_file_extension(file_path)


def create_output_folder(folder_name="拆分结果", clear_old=False):
    return FileUtils.create_output_folder(folder_name, clear_old)


def get_file_name(input_file, idx, ext):
    return FileUtils.get_file_name(input_file, idx, ext)


def detect_encoding(file_path):
    return FileUtils.detect_encoding(file_path)


def count_csv_rows(file_path):
    return FileUtils.count_csv_rows(file_path)


def read_csv_chunks(file_path, batch_size):
    return FileUtils.read_csv_chunks(file_path, batch_size)


def get_excel_data(file_path):
    return FileUtils.get_excel_data(file_path)


def is_valid_file(file_path):
    return FileUtils.is_valid_file(file_path)


def get_file_info(file_path):
    return FileUtils.get_file_info(file_path)
